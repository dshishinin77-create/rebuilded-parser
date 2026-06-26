from ast import Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import re
import os
import copy
import datetime
import json


def smart_load_workbook(file_path):
    return load_workbook(file_path)


def get_cell(ini_cell, row, column):
    return (ini_cell[0] + row, ini_cell[1] + column)


def first_cell_def(cells):
    for key in sorted(list(cells.keys())):
        if cells[key]:
            return key


def first_not_empty(ini_cell, cells, direction):
    if ini_cell not in cells.keys():
        return (None, None)
    else:
        if cells[ini_cell]:
            return (ini_cell, cells[ini_cell])
        else:
            if direction == 'row':
                return first_not_empty(get_cell(ini_cell, 0, 1), cells,
                                       direction)
            else:
                return first_not_empty(get_cell(ini_cell, 1, 0), cells,
                                       direction)


def re_var(consts: list, string: str, stop_words=None):
    string = str(string)
    string = ''.join(string.replace('\n', ''))
    reg = '.*' + '.*'.join(consts) + '.*'
    if stop_words:
        for word in stop_words:
            stop = '.*' + word + '.*'
            if bool(re.match(f'{stop}', string, re.IGNORECASE)):
                return False
    return bool(re.match(f'{reg}', string, re.IGNORECASE))


def header_in_reg(header, position, dict_of_reg):
    if type(dict_of_reg[position][0]) == str:
        _temp_bool = re_var(dict_of_reg[position], header)
    else:
        _temp_bool = re_var(dict_of_reg[position][0], header,
                            dict_of_reg[position][1])
    return _temp_bool


def temp_dict_of_row(row, cells):
    _temp_dict = {}
    cell = (row, 1)
    if cell not in cells:
        return _temp_dict
    value = cells[cell]
    last_cell = list(cells.keys())[-1]

    while cell and cell[1] <= last_cell[1]:
        _temp_dict[cell[1]] = value
        cell = (cell[0], cell[1] + 1)
        cell, value = first_not_empty(cell, cells, 'row')
        if not cell:
            break
    return _temp_dict


def smart_join(massive):
    massive = [i for i in massive if i]
    return '\n'.join(massive)


dict_of_reg_value_FCR = {
    'CR_number': [['change', 'request'], ['init']],
    'Reg_date': ['registr', 'dat'],
    'CR_coordinator': ['contr', 'coord'],
    'CR_number_int': ['init', 'intern'],
    'Organization': [['init', 'organ'], ['organ', 'intern']],
    'Initiator': [['init'], ['organ', 'intern']],
    'Initiator_pos': [['init', 'posit']],
    'Method_justif': [['justification'], ['simple']],
    'Ini_Method_justif': ['just', 'simple', 'meth'],
    'Change_equipment': ['change', 'equip'],
    'Reason_code': ['reason', 'other'],
    'Material_eq': ['equiv', 'repl'],
    'NS': ['nucl', 'saf'],
    'FS': ['fire', 'industr'],
    'IS': ['industr', 'saf'],
    'ES': ['envir'],
    'SS': ['struct', 'geom']
}

dict_of_reg_value_CRD = {
    'CR_number': [['change', 'request'], ['init']],
    'CR_coordinator': ['contr', 'coord'],
    'CR_number_int': ['init', 'intern'],
    'Organization': [['init', 'organ'], ['organ', 'intern']],
    'Initiator': ['init'],
    'Document_type': ['document', 'type'],
    'Method_CR': ['method'],
    'Contract': ['contract'],
    'Impact_cost': ['cost', 'impact'],
    'Schedule': ['schedul'],
    'Comment_nont': ['comment', 'non']
}

dict_of_reg_value_CR = {
    'CR_number': [['change', 'request'], ['init']],
    'Organization': ['init', 'organ'],
    'Responsible': ['responsible', 'eval'],
    'Initiator': [['init'], ['organ', 'intern']],
    'CR_coordinator': ['contr', 'coord'],
    'CR_number_int': ['init', 'intern'],
    'Contract': ['contract'],
    'Impact_cost': ['cost', 'impact'],
    'Schedule': ['schedul'],
    'Comment_nont': ['comment', 'non']
}


def main_func(table_name):
    short_table_name = table_name.split('\\')[-1]
    print(f'{short_table_name} starts.__________')
    warning = 0
    letter_number = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G',
                     8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M',
                     14: 'N', 15: 'O'}
    wb = smart_load_workbook(table_name)
    ws = wb.worksheets[0]
    for sheet in wb.worksheets:
        if sheet.sheet_state == 'visible':
            ws = sheet
            break
    cells = ws._cells
    cells = dict(map(lambda x: (x, cells[x].value), cells))
    cells = {
        key: cells[key] if not re_var(['unnamed'], str(cells[key])) else None
        for key in cells}
    last_cell = list(cells.keys())[-1]
    first_cell = first_cell_def(cells)

    if 'FCR' in table_name.split('/')[-1]:
        dict_of_reg_value_local = dict_of_reg_value_FCR.copy()
        CR_d = {
            'General_information': {
                'CR_reason': None, 'CR_number': None, 'Reg_date': None,
                'CR_coordinator': None,
                'CR_number_int': None, 'Organization': None, 'Initiator': None,
                'Initiator_pos': None,
                'Document_type': None, 'Change_type': None,
                'Activity_type': None, 'Constr_facility': None,
                'Ini_Method_CR': None, 'Method_CR': None,
                'Ini_Method_justif': None, 'Method_justif': None,
                'Change_equipment': None, 'Reason_code': None,
                'Descr_tech_sol': None, 'Final_status': None,
                'Evaluation': {
                    'Material_eq': None, 'REPLACE?': None,
                    'Reject_comment': None, 'Refuse_comment': None,
                    'JD': {}, 'Impact_DDD': None, 'Impact_LDD': None,
                    'Impact_cost': None, 'Schedule': None,
                    'Prompt_req': None, 'NS': None, 'FS': None, 'IS': None,
                    'ES': None, 'SS': None
                }
            },
            'Approval': [],
            'Supp_descr_docs': {},
            'TDD_sets': {},
            'SSC': {}
        }
        section = 'General'
        subsection = False
        _descr_flag = 0

        for row in range(first_cell[0], last_cell[0] + 1):
            if list(filter(
                    lambda x: x[0] == row and x[1] > last_cell[1] and cells[x],
                    cells)):
                print(
                    f'!WARNING! Row: {row}. Value outside of main content field!')
                warning += 1
                continue

            if not any(
                    [cells.get((row, i)) for i in range(1, last_cell[1] + 1)]):
                print(f'!WARNING! Row: {row}. Empty row!')
                if section == 'Concurrence':
                    section = 'Approval'
                continue

            if cells.get((row, 1)):
                if section == 'General' and re_var(['affect', 'ssc'],
                                                   cells[row, 1]):
                    section = 'SSC'
                elif section == 'SSC' and re_var(['list', 'rel', 'doc'],
                                                 cells[row, 1]):
                    section = 'TDD'
                elif section == 'TDD' and re_var(['supp', 'descr'],
                                                 cells[row, 1]):
                    section = 'Sup_doc'
                elif section == 'Sup_doc' and re_var(['evaluat', 'of'],
                                                     cells[row, 1]):
                    section = 'Evaluation'
                elif section == 'Evaluation' and re_var(['linc', 'doc'],
                                                        cells[row, 1]):
                    subsection = 'JD'
                elif section == 'Evaluation' and re_var(['concur', 'sheet'],
                                                        cells[row, 1]):
                    section = 'Concurrence'
                elif section == 'Approval' and re_var(['close'],
                                                      cells[row, 1]):
                    section = 'Close'

                if section == 'SSC' and re_var(['not', 'spec'],
                                               cells[row, 1]): continue
                if section == 'SSC' and re_var(['cod', 'SSC', 'n/a'],
                                               cells[row, 1]): continue
                if section == 'TDD' and re_var(['set', 'which'],
                                               cells[row, 1]): continue
                if section == 'Evaluation' and re_var(['type', 'of', 'change'],
                                                      cells[row, 1]): continue
                if section == 'Evaluation' and re_var(['crit', 'imp'],
                                                      cells[row, 1]): continue
                if section == 'Evaluation' and subsection == 'JD' and re_var(
                    ['file', 'extension'], cells[row, 1]): continue
                if section == 'General' and re_var(['field', 'change', 'fcr'],
                                                   cells[row, 1]): continue
                if section == 'General' and re_var(['change', 'init', 'for'],
                                                   cells[row, 1]): continue

                if section == 'General' and _descr_flag == 1:
                    _descr_flag = 0
                if section == 'Evaluation' and _descr_flag == 1:
                    _descr_flag = 0

                if section == 'Approval' and re_var(['\\*', '\\*\\*'],
                                                    cells[row, 1], stop_words=[
                        '\\*\\*\\*']): continue
                if section == 'Sup_doc' and (
                        re_var(['file', 'ext'], cells[row, 1]) or re_var(
                    ['end', 'init'], cells[row, 1])): continue

                if section in ['TDD', 'SSC', 'Approval']:
                    if section == 'Close' and re_var(['end', 'form'],
                                                     cells[row, 1]):
                        break
                    if re_var(['code'], cells[row, 1]) or re_var(['posit'],
                                                                 cells[
                                                                     row, 1]):
                        _temp_dict = temp_dict_of_row(row, cells)
                        _temp_dict_warning_header = {**_temp_dict}
                        for merged_cells in ws.merged_cells.ranges:
                            if (row, 1) in list(merged_cells.cells):
                                for _row in {_cell[0] for _cell in
                                             list(merged_cells.cells)} - {row}:
                                    _temp_dict_warning_header |= temp_dict_of_row(
                                        _row, cells)
                        extra_value_col = {x[1] for x in cells if
                                           x[0] == row and cells[x]} - set(
                            _temp_dict_warning_header.keys())
                        if len(extra_value_col) > 0:
                            print(
                                f'!WARNING! Row: {row}. Value without header.')
                            warning += 1

                        code = None
                        if section in ['TDD', 'SSC']:
                            code_keys = list(filter(
                                lambda x: re_var(['code'], _temp_dict[x]),
                                _temp_dict))
                            if code_keys: code = cells.get((row, code_keys[0]))
                        elif section == 'Approval':
                            code_keys = list(filter(
                                lambda x: re_var(['posit'], _temp_dict[x]),
                                _temp_dict))
                            if code_keys: code = cells.get((row, code_keys[0]))

                        if not code:
                            print(
                                f'    !WARNING! Row: {row}. KKS or Position is NONE!')
                            continue

                        if section == 'SSC':
                            CR_d['SSC'][code] = {}
                        elif section == 'Approval':
                            CR_d['Approval'].append({'Position': code})
                        else:
                            if code not in CR_d['TDD_sets'].keys():
                                CR_d['TDD_sets'][code] = {}

                        for column_number in _temp_dict:
                            insert_value = cells.get((row, column_number))
                            if section in 'TDD' and insert_value:
                                if re_var(['set', 'revis'],
                                          _temp_dict[column_number]):
                                    if '_' in str(insert_value):
                                        CR_d['TDD_sets'][code]['Revision'] = \
                                        str(insert_value).split('_')[0]
                                        CR_d['TDD_sets'][code]['Version'] = \
                                        str(insert_value).split('_')[1]
                                    else:
                                        CR_d['TDD_sets'][code][
                                            'Revision'] = insert_value
                                        CR_d['TDD_sets'][code]['Version'] = 0
                                if re_var(['engin', 'cod'],
                                          _temp_dict[column_number]):
                                    doc_code = insert_value
                                    if 'Documents' not in CR_d['TDD_sets'][
                                        code].keys():
                                        CR_d['TDD_sets'][code]['Documents'] = [
                                            {'Code': doc_code}]
                                    if doc_code not in list(
                                            map(lambda x: x['Code'],
                                                CR_d['TDD_sets'][code][
                                                    'Documents'])):
                                        CR_d['TDD_sets'][code][
                                            'Documents'].append(
                                            {'Code': doc_code})
                                if re_var(['engin', 'name'],
                                          _temp_dict[column_number]):
                                    index_of_doc_code = \
                                    [index for index, value in enumerate(
                                        CR_d['TDD_sets'][code]['Documents']) if
                                     value['Code'] == doc_code][0]
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code][
                                        'Name'] = insert_value
                                if re_var(['revis', 'ED'],
                                          _temp_dict[column_number]) or re_var(
                                        ['ED', 'revis'],
                                        _temp_dict[column_number]):
                                    if '_' in str(insert_value):
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Revision'] = \
                                        str(insert_value).split('_')[0]
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Version'] = \
                                        str(insert_value).split('_')[1]
                                    else:
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code][
                                            'Revision'] = insert_value
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Version'] = 0
                                if re_var(['sheets'],
                                          _temp_dict[column_number]):
                                    sheets = str(insert_value)
                                    if 'Sheets' not in CR_d['TDD_sets'][code][
                                        'Documents'][index_of_doc_code]:
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Sheets'] = {}
                                    if sheets not in CR_d['TDD_sets'][code][
                                        'Documents'][index_of_doc_code][
                                        'Sheets'].keys():
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Sheets'][
                                            sheets] = {}
                                if re_var(['chang', 'amx'],
                                          _temp_dict[column_number]):
                                    if 'AMX' not in CR_d['TDD_sets'][code][
                                        'Documents'][index_of_doc_code][
                                        'Sheets'][sheets].keys():
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Sheets'][
                                            sheets]['AMX'] = insert_value
                                    else:
                                        if insert_value not in \
                                                CR_d['TDD_sets'][code][
                                                    'Documents'][
                                                    index_of_doc_code][
                                                    'Sheets'][sheets]['AMX']:
                                            CR_d['TDD_sets'][code][
                                                'Documents'][
                                                index_of_doc_code]['Sheets'][
                                                sheets][
                                                'AMX'] += f'\n{insert_value}'
                                if re_var(['description'],
                                          _temp_dict[column_number]):
                                    if 'Description' not in \
                                            CR_d['TDD_sets'][code][
                                                'Documents'][
                                                index_of_doc_code]['Sheets'][
                                                sheets].keys():
                                        CR_d['TDD_sets'][code]['Documents'][
                                            index_of_doc_code]['Sheets'][
                                            sheets]['Description'] = [
                                            insert_value]
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code]['Sheets'][sheets][
                                        'Description'].append(insert_value)
                            if section == 'SSC':
                                if re_var(['system', 'code'],
                                          _temp_dict[column_number]):
                                    CR_d['SSC'][code][
                                        'Sys_code'] = insert_value
                                elif re_var(['component'],
                                            _temp_dict[column_number]):
                                    CR_d['SSC'][code][
                                        'Component'] = insert_value
                            if section == 'Approval':
                                if re_var(['person'],
                                          _temp_dict[column_number]):
                                    CR_d['Approval'][-1]['Name'] = insert_value
                                elif re_var(['date'],
                                            _temp_dict[column_number]):
                                    CR_d['Approval'][-1]['Date'] = insert_value

                if section in ['General', 'Concurrence', 'Close'] or (
                        section == 'Evaluation' and subsection != 'JD'):
                    if section == 'General' and re_var(
                            ['descrip', 'engin', 'change'],
                            cells.get((row, 1))):
                        CR_d['General_information'][
                            'Descr_tech_sol'] = cells.get((row + 1, 1))
                        _descr_flag = 1
                    if section == 'Evaluation' and re_var(
                            ['comment', 'reason', 'reject'],
                            cells.get((row, 1))):
                        CR_d['General_information']['Evaluation'][
                            'Reject_comment'] = cells.get((row + 1, 1))
                        _descr_flag = 1
                    if section == 'Evaluation' and re_var(
                            ['comment', 'reason', 'refus'],
                            cells.get((row, 1))):
                        CR_d['General_information']['Evaluation'][
                            'Refuse_comment'] = cells.get((row + 1, 1))
                        _descr_flag = 1

                    _temp_dict = temp_dict_of_row(row, cells)
                    if len(_temp_dict.keys()) % 2 != 0:
                        if len(_temp_dict.keys()) == 1:
                            _temp_dict[list(_temp_dict.keys())[0] + 1] = None
                        else:
                            if not any(map(lambda x: header_in_reg(
                                    _temp_dict[list(_temp_dict.keys())[-1]], x,
                                    dict_of_reg_value_FCR),
                                           dict_of_reg_value_FCR.keys())):
                                if not any(map(lambda x: header_in_reg(
                                        _temp_dict[
                                            list(_temp_dict.keys())[-2]], x,
                                        dict_of_reg_value_FCR),
                                               dict_of_reg_value_FCR.keys())):
                                    _temp_dict.pop(list(_temp_dict.keys())[-1])

                    _temp_temp_dict = {}
                    flag = False
                    key = 0
                    for item in list(sorted(_temp_dict.keys())):
                        if any(map(lambda x: header_in_reg(_temp_dict[item], x,
                                                           dict_of_reg_value_FCR),
                                   dict_of_reg_value_FCR.keys())):
                            if flag == True:
                                _temp_temp_dict[key + 1] = None
                            _temp_temp_dict[item] = _temp_dict[item]
                            flag = True
                            key = item
                        else:
                            if flag == True:
                                _temp_temp_dict[item] = _temp_dict[item]
                                flag = False
                    if flag == True:
                        _temp_temp_dict[key + 1] = None
                    _temp_dict = _temp_temp_dict
                    sorted_dict_keys = sorted(_temp_dict.keys())
                    _temp_dict = {
                        _temp_dict[k]: _temp_dict[sorted_dict_keys[i + 1]] for
                        i, k in enumerate(sorted_dict_keys) if i % 2 == 0}

                    for header in _temp_dict.keys():
                        for position in list(dict_of_reg_value_local.keys()):
                            if header_in_reg(header, position,
                                             dict_of_reg_value_FCR):
                                if section == 'Evaluation':
                                    CR_d['General_information']['Evaluation'][
                                        position] = _temp_dict[header]
                                else:
                                    CR_d['General_information'][position] = \
                                    _temp_dict[header]
                                dict_of_reg_value_local.pop(position)
                                break

                    if section == 'Sup_doc' or (
                            section == 'Evaluation' and subsection == 'JD'):
                        if section == 'Sup_doc':
                            if str(cells.get((row, 1))) in CR_d[
                                'Supp_descr_docs'].keys():
                                CR_d['Supp_descr_docs'][str(cells[row, 1])][
                                    'Title'].append(
                                    first_not_empty((row, 2), cells, 'row')[1])
                            else:
                                curr_cell, curr_cell_value = first_not_empty(
                                    (row, 2), cells, 'row')
                                CR_d['Supp_descr_docs'][
                                    str(cells.get((row, 1)))] = {
                                    'Title': [curr_cell_value]}
                            curr_cell, \
                            CR_d['General_information']['Evaluation']['JD'][
                                str(cells.get((row, 1)))] = first_not_empty(
                                (row, 2), cells, 'row')
                            if re_var(['type', 'change'], cells.get((row, 1))):
                                subsection = False

            if not cells.get((row, 1)) and section == 'Concurrence':
                section = 'Approval'

    elif 'CR.D-' in table_name.split('/')[-1]:
        dict_of_reg_value_local = dict_of_reg_value_CRD.copy()
        CR_d = {
            'General_information': {'CR_number': None, 'CR_coordinator': None,
                                    'CR_number_int': None,
                                    'Organization': None, 'Initiator': None,
                                    'TDD_influece': None,
                                    'Document_type': None, 'CR_reason': None,
                                    'Descr_tech_sol': None,
                                    'NSC_category': None, 'Impact_1_2_3': None,
                                    'Impact_DSA': None, 'Method_CR': None,
                                    'Comment': None, 'Contract': None,
                                    'Impact_cost': None, 'Schedule': None,
                                    'Comment_nont': None}, 'Confirmation': {},
            'Approval': {}, 'Supp_descr_docs': {}, 'TDD': {}, 'SSC': {}}
        section = 'General'
        for row in range(first_cell[0], last_cell[0] + 1):
            if list(filter(
                    lambda x: x[0] == row and x[1] > last_cell[1] and cells[x],
                    cells)):
                print(
                    f'    !WARNING! Row: {row}. Value outside of main content field!')
                warning += 1
                continue
            if not any(
                    [cells.get((row, i)) for i in range(1, last_cell[1] + 1)]):
                print(f'    !WARNING! Row: {row}. Empty row!')
                continue

            if cells.get((row, 1)):
                if section == 'General' and re_var(['supporting', 'descr'],
                                                   cells[row, 1]):
                    section = 'Sup_doc'
                elif section == 'Sup_doc' and re_var(['impact', 'init', 'TDD'],
                                                     cells[row, 1]):
                    section = 'TDD'
                elif section == 'TDD' and re_var(['impact', 'TDD'],
                                                 cells[row, 1]):
                    section = 'Other_TDD'
                elif section == 'Other_TDD' and re_var(['ssc'], cells[row, 1]):
                    section = 'SSC'
                elif section == 'SSC' and re_var(['final', 'eval'],
                                                 cells[row, 1]):
                    section = 'Final'
                elif section == 'Final' and (
                        re_var(['non-tech'], cells[row, 1]) or re_var(
                        ['nontech'], cells[row, 1])):
                    section = 'Nontech'
                elif section == 'Nontech' and re_var(['confirmat'],
                                                     cells[row, 1]):
                    section = 'Confirmation'
                elif section == 'Confirmation' and re_var(['approv'],
                                                          cells[row, 1]):
                    section = 'Approval'

                if section == 'Sup_doc' and re_var(['file', 'extension'],
                                                   cells[row, 1]): continue
                if section == 'General' and re_var(['general', 'inform'],
                                                   cells[row, 1]): continue

                if section in ['TDD', 'Other_TDD', 'SSC', 'Confirmation',
                               'Approval']:
                    if section == 'Approval' and list(
                            filter(lambda x: x.size['columns'] == last_cell[
                                1] and x.start_cell.row == row,
                                   list(ws.merged_cells.ranges))) or \
                            str(cells[row, 1])[0] == '*':
                        break
                    else:
                        if re_var(['code'], cells[row, 1]) or re_var(['organ'],
                                                                     cells[
                                                                         row, 1]) or re_var(
                                ['posit'], cells[row, 1]):
                            _temp_dict = temp_dict_of_row(row, cells)
                            _temp_dict_warning_header = {**_temp_dict}
                            for merged_cells in ws.merged_cells.ranges:
                                if (row, 1) in list(merged_cells.cells):
                                    for _row in {_cell[0] for _cell in
                                                 list(merged_cells.cells)} - {
                                                    row}:
                                        _temp_dict_warning_header |= temp_dict_of_row(
                                            _row, cells)
                        else:
                            extra_value_col = {x[1] for x in cells if
                                               x[0] == row and cells[x]} - set(
                                _temp_dict_warning_header.keys())
                            if len(extra_value_col) > 0:
                                print(
                                    f'    !WARNING! Row: {row}. Value without header.')
                                warning += 1

                            if section in ['TDD', 'Other_TDD', 'SSC']:
                                code_keys = list(filter(
                                    lambda x: re_var(['code'], _temp_dict[x]),
                                    _temp_dict))
                                code = code_keys[0] if code_keys else None
                            elif section == 'Confirmation':
                                code_keys = list(filter(
                                    lambda x: re_var(['organ'], _temp_dict[x]),
                                    _temp_dict))
                                code = code_keys[0] if code_keys else None
                            elif section == 'Approval':
                                code_keys = list(filter(
                                    lambda x: re_var(['posit'], _temp_dict[x]),
                                    _temp_dict))
                                code = code_keys[0] if code_keys else None

                            code = cells.get((row, code)) if code else None
                            if not code:
                                print(
                                    f'    !WARNING! Row: {row}. KKS is NONE!')
                                continue

                            if section == 'SSC':
                                CR_d['SSC'][code] = {}
                            elif section == 'Confirmation':
                                if code not in CR_d['Confirmation'].keys():
                                    CR_d['Confirmation'][code] = {}
                            elif section == 'Approval':
                                CR_d['Approval'][code] = {}
                            else:
                                if code not in CR_d['TDD'].keys():
                                    CR_d['TDD'][code] = {}

                            for column_number in _temp_dict:
                                insert_value = cells.get((row, column_number))
                                if section in ['TDD',
                                               'Other_TDD'] and insert_value:
                                    if re_var(['organ'],
                                              _temp_dict[column_number]):
                                        CR_d['TDD'][code][
                                            'Organization'] = insert_value
                                    elif re_var(['new', 'revision'],
                                                _temp_dict[column_number]):
                                        _temp_bool = True if 'es' in str(
                                            insert_value).lower() else False
                                        CR_d['TDD'][code][
                                            'New_rev_req'] = _temp_bool
                                    elif re_var(['revision'],
                                                _temp_dict[column_number]):
                                        if '_' in str(insert_value):
                                            CR_d['TDD'][code]['Revision'] = \
                                            str(insert_value).split('_')[0]
                                            CR_d['TDD'][code]['Version'] = \
                                            str(insert_value).split('_')[1]
                                        else:
                                            CR_d['TDD'][code][
                                                'Revision'] = insert_value
                                            CR_d['TDD'][code]['Version'] = 0
                                    elif re_var(['name'],
                                                _temp_dict[column_number]):
                                        CR_d['TDD'][code][
                                            'Name'] = insert_value
                                    elif re_var(['state'],
                                                _temp_dict[column_number]):
                                        CR_d['TDD'][code][
                                            'Status'] = insert_value
                                    elif re_var(['description'],
                                                _temp_dict[column_number]):
                                        if 'Description' not in CR_d['TDD'][
                                            code].keys():
                                            CR_d['TDD'][code][
                                                'Description'] = [insert_value]
                                        else:
                                            CR_d['TDD'][code][
                                                'Description'].append(
                                                insert_value)
                                    elif re_var(['impact'],
                                                _temp_dict[column_number]):
                                        list_of_factors = ''
                                        for factor in range(5):
                                            _temp_bin = '0' if re_var(['no'],
                                                                      cells.get(
                                                                          (row,
                                                                           column_number + factor))) else '1'
                                            list_of_factors += _temp_bin
                                        CR_d['TDD'][code]['Impact'] = int(
                                            list_of_factors, 2)
                                elif section == 'SSC':
                                    if re_var(['name'],
                                              _temp_dict[column_number]):
                                        CR_d['SSC'][code][
                                            'Name'] = insert_value
                                    elif re_var(['description'],
                                                _temp_dict[column_number]):
                                        CR_d['SSC'][code][
                                            'Description'] = insert_value
                                elif section == 'Confirmation':
                                    if re_var(['posit'],
                                              _temp_dict[column_number]):
                                        position = insert_value
                                    elif re_var(['resp', 'pers'],
                                                _temp_dict[column_number]):
                                        name = insert_value
                                        CR_d['Confirmation'][code][name] = {
                                            'Position': None, 'Date': None}
                                    elif re_var(['date'],
                                                _temp_dict[column_number]):
                                        date = insert_value
                                        CR_d['Confirmation'][code][name][
                                            'Position'] = position
                                        CR_d['Confirmation'][code][name][
                                            'Date'] = date
                                elif section == 'Approval':
                                    if re_var(['person'],
                                              _temp_dict[column_number]):
                                        CR_d['Approval'][code][
                                            'person'] = insert_value
                                    elif re_var(['date'],
                                                _temp_dict[column_number]):
                                        CR_d['Approval'][code][
                                            'date'] = insert_value

                if section in ['General', 'Final', 'Nontech']:
                    _temp_dict = temp_dict_of_row(row, cells)
                    if len(_temp_dict.keys()) % 2 != 0:
                        if len(_temp_dict.keys()) == 1:
                            _temp_dict[list(_temp_dict.keys())[0] + 1] = None
                        else:
                            flag = False
                            key = 0
                            for item in list(sorted(_temp_dict.keys())):
                                for position in dict_of_reg_value_CRD.keys():
                                    if header_in_reg(_temp_dict[item],
                                                     position,
                                                     dict_of_reg_value_CRD) and flag == False:
                                        flag = True
                                        key = item
                                    elif header_in_reg(_temp_dict[item],
                                                       position,
                                                       dict_of_reg_value_CRD) and flag == True:
                                        _temp_dict[key + 1] = None
                    sorted_dict_keys = sorted(_temp_dict.keys())
                    _temp_dict = {
                        _temp_dict[k]: _temp_dict[sorted_dict_keys[i + 1]] for
                        i, k in enumerate(sorted_dict_keys) if i % 2 == 0}
                    for header in _temp_dict.keys():
                        for position in list(dict_of_reg_value_local.keys()):
                            if header_in_reg(header, position,
                                             dict_of_reg_value_CRD):
                                CR_d['General_information'][position] = \
                                _temp_dict[header]
                                dict_of_reg_value_local.pop(position)
                                break
                    if section == 'Sup_doc':
                        curr_cell, CR_d['Supp_descr_docs'][
                            str(cells.get((row, 1)))] = first_not_empty(
                            (row, 2), cells, 'row')

    elif 'CR' in table_name.split('/')[-1]:
        dict_of_reg_value_local = dict_of_reg_value_CR.copy()
        CR_d = {
            'General_information': {'CR_number': None, 'Organization': None,
                                    'Initiator': None, 'Responsible': None,
                                    'Impact_DSA': None, 'CR_coordinator': None,
                                    'CR_number_int': None, 'Contract': None,
                                    'Impact_cost': None, 'Schedule': None,
                                    'Comment_nont': None}, 'Confirmation': {},
            'Approval': {}, 'Configur': {}, 'TDD': {}, 'SSC': {}}
        section = 'General'
        for row in range(first_cell[0], last_cell[0] + 1):
            if list(filter(
                    lambda x: x[0] == row and x[1] > last_cell[1] and cells[x],
                    cells)):
                print(
                    f'    !WARNING! Row: {row}. Value outside of main content field!')
                warning += 1
            else:
                if not any([cells.get((row, i)) for i in
                            range(1, last_cell[1] + 1)]):
                    print(f'    !WARNING! Row: {row}. Empty row!')
                    continue
                else:
                    if section == 'General' and re_var(['init', 'item'],
                                                       cells[row, 1]):
                        section = 'TDD'
                    elif section == 'TDD' and re_var(['conf', 'item'],
                                                     cells[row, 1]):
                        section = 'Configur'
                    elif section == 'Configur' and re_var(['affect', 'syst'],
                                                          cells[row, 1]):
                        section = 'SSC'
                    elif section == 'SSC' and re_var(['confirmat'],
                                                     cells[row, 1]):
                        section = 'Confirmation'
                    elif section == 'Confirmation' and (
                            re_var(['non-tech'], cells[row, 1]) or re_var(
                            ['nontech'], cells[row, 1])):
                        section = 'Nontech'
                    elif section == 'Nontech' and re_var(['approv'],
                                                         cells[row, 1]):
                        section = 'Approval'

                    if section in ['TDD', 'Configur', 'SSC', 'Confirmation',
                                   'Approval'] and (
                            re_var(['code'], cells.get((row, 1))) or re_var(
                            ['posit'], cells.get((row, 1)))):
                        _temp_dict = temp_dict_of_row(row, cells)
                        _temp_dict_warning_header = {**_temp_dict}
                        for merged_cells in ws.merged_cells.ranges:
                            if (row, 1) in list(merged_cells.cells):
                                for _row in {_cell[0] for _cell in
                                             list(merged_cells.cells)} - {row}:
                                    _temp_dict_warning_header |= temp_dict_of_row(
                                        _row, cells)
                        if section == 'SSC':
                            not_empty = \
                            first_not_empty((row, 2), cells, 'row')[0]
                            not_empty = (not_empty[0], not_empty[1] + 1)
                            cell_of_DSA = \
                            first_not_empty(not_empty, cells, 'row')[1]
                            CR_d['General_information'][
                                'Impact_DSA'] = cell_of_DSA
                    elif section in ['TDD', 'Configur', 'SSC', 'Confirmation',
                                     'Approval']:
                        extra_value_col = {x[1] for x in cells if
                                           x[0] == row and cells[x]} - set(
                            _temp_dict_warning_header.keys())
                        if len(extra_value_col) > 0:
                            print(
                                f'    !WARNING! Row: {row}. Value without header.')
                            warning += 1

                        code = None
                        if section in ['TDD', 'Configur', 'SSC']:
                            code_keys = list(filter(
                                lambda x: re_var(['code'], _temp_dict[x]),
                                _temp_dict))
                            if code_keys: code = code_keys[0]
                        elif section in ['Approval', 'Confirmation']:
                            code_keys = list(filter(
                                lambda x: re_var(['posit'], _temp_dict[x]),
                                _temp_dict))
                            if code_keys: code = code_keys[0]

                        if cells.get((row, code)):
                            code = cells[(row, code)]
                        else:
                            print(f'    !WARNING! Row: {row}. KKS is NONE!')
                            continue

                        if section == 'SSC':
                            CR_d['SSC'][code] = {}
                        elif section == 'Confirmation':
                            if code not in CR_d['Confirmation'].keys():
                                CR_d['Confirmation'][code] = {}
                        elif section == 'Approval':
                            CR_d['Approval'][code] = {}
                        elif section == 'Configur':
                            CR_d['Configur'][code] = {}
                        else:
                            if code not in list(CR_d['TDD'].keys()):
                                CR_d['TDD'][code] = {}

                        for column_number in _temp_dict:
                            insert_value = cells.get((row, column_number))
                            if section in ['TDD', 'Configur'] and insert_value:
                                if re_var(['revision'],
                                          _temp_dict[column_number]):
                                    if '_' in str(insert_value):
                                        CR_d[section][code]['Revision'] = \
                                        str(insert_value).split('_')[0]
                                        CR_d[section][code]['Version'] = \
                                        str(insert_value).split('_')[1]
                                    else:
                                        CR_d[section][code][
                                            'Revision'] = insert_value
                                        CR_d[section][code]['Version'] = 0
                                elif re_var(['name'],
                                            _temp_dict[column_number]):
                                    CR_d[section][code]['Name'] = insert_value
                                elif re_var(['state'],
                                            _temp_dict[column_number]):
                                    CR_d[section][code][
                                        'Status'] = insert_value
                                elif re_var(['description'],
                                            _temp_dict[column_number]):
                                    if 'Description' not in list(
                                            CR_d[section][code].keys()):
                                        CR_d[section][code][
                                            'Description'] = insert_value
                                    else:
                                        CR_d[section][code][
                                            'Description'] += insert_value
                                        print(
                                            f'    !WARNING! Row: {row}. TDD description w/o code.')
                                        warning += 1
                                elif re_var(['reason'],
                                            _temp_dict[column_number]):
                                    if 'Merged' in str(type(
                                            ws._cells[row, column_number])):
                                        merged_cells = list(filter(lambda
                                                                       x: x.min_col <= column_number <= x.max_col and x.min_row <= row <= x.max_row,
                                                                   list(
                                                                       ws.merged_cells.ranges)))[
                                            0]
                                        insert_value = cells[
                                            merged_cells.min_row, merged_cells.min_col]
                                    CR_d[section][code][
                                        'Reason'] = insert_value
                                elif re_var(['eval', 'imp'],
                                            _temp_dict[column_number]):
                                    CR_d[section][code][
                                        'Imp_eval'] = insert_value
                            elif section == 'SSC':
                                if re_var(['description'],
                                          _temp_dict[column_number]):
                                    CR_d['SSC'][code][
                                        'Description'] = insert_value
                            elif section == 'Confirmation':
                                if re_var(['posit'],
                                          _temp_dict[column_number]):
                                    position = insert_value
                                elif re_var(['resp', 'pers'],
                                            _temp_dict[column_number]):
                                    name = insert_value
                                    CR_d['Confirmation'][code][name] = {
                                        'Position': None, 'Date': None}
                                elif re_var(['date'],
                                            _temp_dict[column_number]):
                                    date = insert_value
                                    CR_d['Confirmation'][code][name][
                                        'Position'] = position
                                    CR_d['Confirmation'][code][name][
                                        'Date'] = date
                            elif section == 'Approval':
                                if re_var(['person'],
                                          _temp_dict[column_number]):
                                    CR_d['Approval'][code][
                                        'person'] = insert_value
                                elif re_var(['date'],
                                            _temp_dict[column_number]):
                                    CR_d['Approval'][code][
                                        'date'] = insert_value

                    elif section in ['General', 'Nontech']:
                        _temp_dict = temp_dict_of_row(row, cells)
                        if len(_temp_dict.keys()) % 2 != 0:
                            if len(_temp_dict.keys()) == 1:
                                _temp_dict[
                                    list(_temp_dict.keys())[0] + 1] = None
                            else:
                                flag = False
                                key = 0
                                for item in list(sorted(_temp_dict.keys())):
                                    for position in dict_of_reg_value_CR.keys():
                                        if header_in_reg(_temp_dict[item],
                                                         position,
                                                         dict_of_reg_value_CR) and flag == False:
                                            flag = True
                                            key = item
                                        elif header_in_reg(_temp_dict[item],
                                                           position,
                                                           dict_of_reg_value_CR) and flag == True:
                                            _temp_dict[key + 1] = None
                        sorted_dict_keys = sorted(_temp_dict.keys())
                        _temp_dict = {
                            _temp_dict[k]: _temp_dict[sorted_dict_keys[i + 1]]
                            for i, k in enumerate(sorted_dict_keys) if
                            i % 2 == 0}
                        for header in _temp_dict.keys():
                            for position in list(
                                    dict_of_reg_value_local.keys()):
                                if header_in_reg(header, position,
                                                 dict_of_reg_value_CR):
                                    CR_d['General_information'][position] = \
                                    _temp_dict[header]
                                    dict_of_reg_value_local.pop(position)
                                    break
    if warning:
        print(
            f'{short_table_name}. {warning} warnings were found! JSON will not be done!')
    else:
        print(f'{short_table_name} finished. JSON will be done.')
        return CR_d


NS_template = {
    'File_name': str, 'Change_request_No': str, 'Reg_date': str,
    'Contr_change_coord': str,
    'Type_of_changes': str, 'Constr_facility': str, 'Type_of_changed_doc': str,
    'E-log?': bool,
    'Ini_org': str, 'Change_ini': str, 'Ini_internal_CR': str,
    'Cod_of_reason': str,
    'Other_reason': bool, 'Descr_of_change': str, 'Rel_to_prev_CR': str,
    'Approval_method': str,
    'Just_simple': str, 'Signature_list': [], 'TDD': [], 'SSC': [],
    'Support_files': []
}


def dicts_normalization(original_dict_name):
    normalized_dict = copy.deepcopy(NS_template)
    original_dict = result[original_dict_name].copy()
    normalized_dict['File_name'] = original_dict_name

    if 'CR-' in original_dict_name:
        normalized_dict['Change_request_No'] = \
        original_dict['General_information']['CR_number']
        normalized_dict['Reg_date'] = normalized_dict['Contr_change_coord'] = \
        normalized_dict['Type_of_changes'] = normalized_dict[
            'Constr_facility'] = normalized_dict['Type_of_changed_doc'] = \
        normalized_dict['E-log?'] = None
        normalized_dict['Ini_org'] = original_dict['General_information'][
            'Organization']
        normalized_dict['Change_ini'] = original_dict['General_information'][
            'Initiator']
        normalized_dict['Ini_internal_CR'] = normalized_dict[
            'Cod_of_reason'] = None
        normalized_dict['Other_reason'] = '\n'.join(
            filter(lambda x: x is not None, list(
                map(lambda x: original_dict['TDD'][x]['Reason'],
                    original_dict['TDD']))))
        normalized_dict['Descr_of_change'] = normalized_dict[
            'Rel_to_prev_CR'] = normalized_dict['Approval_method'] = \
        normalized_dict['Just_simple'] = normalized_dict['NSC_category'] = None
        normalized_dict['Impact_direct_123'] = True if original_dict[
            'SSC'] else False
        normalized_dict['Impact_DSA'] = original_dict['General_information'][
            'Impact_DSA']
        normalized_dict['Impact_structural_geom'] = normalized_dict[
            'Impact_nucl'] = normalized_dict['Impact_fire'] = normalized_dict[
            'Impact_industrial'] = normalized_dict['Impact_environment'] = \
        normalized_dict['Impact_TDD'] = normalized_dict['Impact_lic_doc'] = \
        normalized_dict['Prompt_TDD?'] = normalized_dict[
            'Material_equivalent?'] = normalized_dict[
            'Comments_eng_eval'] = None
        normalized_dict['Impact_contract'] = \
        original_dict['General_information']['Contract']
        normalized_dict['Impact_cost'] = original_dict['General_information'][
            'Impact_cost']
        normalized_dict['Impact_schedule'] = \
        original_dict['General_information']['Schedule']
        normalized_dict['Comments_nont_ass'] = \
        original_dict['General_information']['Comment_nont']

        for key in original_dict['Confirmation']:
            _temp_dict = {}
            _temp_name = list(original_dict['Confirmation'][key].keys())[0]
            _temp_dict['Role'] = None
            _temp_dict['Position'] = key
            _temp_dict['Name_sur'] = _temp_name
            _temp_dict['FMV_number'] = _temp_dict['HAEA_reg'] = None
            _temp_dict['Date'] = \
            original_dict['Confirmation'][key][_temp_name]['Date']
            normalized_dict['Signature_list'].append(_temp_dict)

        for key in original_dict['Approval']:
            _temp_dict = {}
            _temp_dict['Role'] = None
            _temp_dict['Position'] = key
            _temp_dict['Name_sur'] = original_dict['Approval'][key]['person']
            _temp_dict['FMV_number'] = _temp_dict['HAEA_reg'] = None
            _temp_dict['Date'] = original_dict['Approval'][key]['date']
            normalized_dict['Signature_list'].append(_temp_dict)

        for key in original_dict['TDD']:
            _temp_dict = {}
            _temp_dict['Type'] = 'TDD'
            _temp_dict['Set_code'] = _temp_dict['Set_name'] = _temp_dict[
                'Set_rev'] = _temp_dict['Ser_version'] = _temp_dict[
                'Set_status'] = None
            _temp_dict['ED_code'] = key
            _temp_dict['ED_name'] = original_dict['TDD'][key]['Name']
            _temp_dict['ED_rev'] = original_dict['TDD'][key]['Revision']
            _temp_dict['ED_version'] = original_dict['TDD'][key]['Version']
            _temp_dict['ED_status'] = original_dict['TDD'][key]['Status']
            _temp_dict['Changed_sheets'] = _temp_dict['AMX_AM'] = None
            _temp_dict['Descr_of_change'] = original_dict['TDD'][key][
                'Description']
            _temp_dict['New_rev?'] = None
            normalized_dict['TDD'].append(_temp_dict)

        for key in original_dict['Configur']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Configur'
            _temp_dict['Set_code'] = _temp_dict['Set_name'] = _temp_dict[
                'Set_rev'] = _temp_dict['Ser_version'] = _temp_dict[
                'Set_status'] = None
            _temp_dict['ED_code'] = key
            _temp_dict['ED_name'] = original_dict['Configur'][key]['Name']
            _temp_dict['ED_rev'] = original_dict['Configur'][key]['Revision']
            _temp_dict['ED_version'] = original_dict['Configur'][key][
                'Version']
            _temp_dict['ED_status'] = original_dict['Configur'][key]['Status']
            _temp_dict['Changed_sheets'] = _temp_dict['AMX_AM'] = None
            _temp_dict['Descr_of_change'] = original_dict['Configur'][key][
                'Description']
            _temp_dict['New_rev?'] = original_dict['Configur'][key]['Imp_eval']
            normalized_dict['TDD'].append(_temp_dict)

        for key in original_dict['SSC']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Changed'
            _temp_dict['List_SSC'] = key
            _temp_dict['Name_KKS'] = None
            _temp_dict['Descr_of_change_KKS'] = original_dict['SSC'][key][
                'Description']
            normalized_dict['SSC'].append(_temp_dict)

    elif 'FCR' in original_dict_name:
        normalized_dict['Change_request_No'] = \
        original_dict['General_information']['CR_number']
        normalized_dict['Reg_date'] = original_dict['General_information'][
            'Reg_date']
        normalized_dict['Contr_change_coord'] = \
        original_dict['General_information']['CR_coordinator']
        normalized_dict['Type_of_changes'] = \
        original_dict['General_information']['Change_type']
        normalized_dict['Constr_facility'] = \
        original_dict['General_information']['Constr_facility']
        normalized_dict['Type_of_changed_doc'] = \
        original_dict['General_information']['Document_type']
        normalized_dict['E-log?'] = None
        normalized_dict['Ini_org'] = original_dict['General_information'][
            'Organization']
        normalized_dict['Change_ini'] = original_dict['General_information'][
            'Initiator']
        normalized_dict['Ini_internal_CR'] = \
        original_dict['General_information']['CR_number_int']
        normalized_dict['Cod_of_reason'] = \
        original_dict['General_information']['Reason_code']
        normalized_dict['Other_reason'] = original_dict['General_information'][
            'CR_reason']
        normalized_dict['Descr_of_change'] = \
        original_dict['General_information']['Descr_tech_sol']
        normalized_dict['Rel_to_prev_CR'] = None
        normalized_dict['Approval_method'] = \
        original_dict['General_information']['Method_CR']
        normalized_dict['Just_simple'] = original_dict['General_information'][
            'Ini_Method_justif']
        normalized_dict['NSC_category'] = None
        normalized_dict['Impact_direct_123'] = True if original_dict[
            'SSC'] else False
        normalized_dict['Impact_DSA'] = None
        normalized_dict['Impact_structural_geom'] = \
        original_dict['General_information']['Evaluation']['SS']
        normalized_dict['Impact_nucl'] = \
        original_dict['General_information']['Evaluation']['NS']
        normalized_dict['Impact_fire'] = \
        original_dict['General_information']['Evaluation']['FS']
        normalized_dict['Impact_industrial'] = \
        original_dict['General_information']['Evaluation']['IS']
        normalized_dict['Impact_environment'] = \
        original_dict['General_information']['Evaluation']['ES']
        normalized_dict['Impact_TDD'] = \
        original_dict['General_information']['Evaluation']['Impact_DDD']
        normalized_dict['Impact_lic_doc'] = \
        original_dict['General_information']['Evaluation']['Impact_LDD']
        normalized_dict['Prompt_TDD?'] = \
        original_dict['General_information']['Evaluation']['Prompt_req']
        normalized_dict['Material_equivalent?'] = \
        original_dict['General_information']['Evaluation']['Material_eq']
        normalized_dict['Comments_eng_eval'] = smart_join([original_dict[
                                                               'General_information'][
                                                               'Evaluation'][
                                                               'Reject_comment'],
                                                           original_dict[
                                                               'General_information'][
                                                               'Evaluation'][
                                                               'Refuse_comment']])
        normalized_dict['Impact_contract'] = None
        normalized_dict['Impact_cost'] = \
        original_dict['General_information']['Evaluation']['Impact_cost']
        normalized_dict['Impact_schedule'] = \
        original_dict['General_information']['Evaluation']['Schedule']
        normalized_dict['Comments_nont_ass'] = None

        for key in original_dict['Approval']:
            _temp_dict = {}
            _temp_dict['Role'] = None
            _temp_dict['Position'] = key['Position']
            _temp_dict['Name_sur'] = key['Name']
            _temp_dict['FMV_number'] = _temp_dict['HAEA_reg'] = None
            _temp_dict['Date'] = key['Date']
            normalized_dict['Signature_list'].append(_temp_dict)

        for key in original_dict['TDD_sets']:
            for document in original_dict['TDD_sets'][key]['Documents']:
                _temp_dict = {}
                _temp_dict['Type'] = 'TDD'
                _temp_dict['Set_code'] = key
                _temp_dict['Set_name'] = None
                _temp_dict['Set_rev'] = original_dict['TDD_sets'][key][
                    'Revision']
                _temp_dict['Ser_version'] = original_dict['TDD_sets'][key][
                    'Version']
                _temp_dict['Set_status'] = None
                _temp_dict['ED_code'] = document
                _temp_dict['ED_name'] = \
                original_dict['TDD_sets'][key]['Documents'][document]['Name']
                _temp_dict['ED_rev'] = \
                original_dict['TDD_sets'][key]['Documents'][document][
                    'Revision']
                _temp_dict['ED_version'] = \
                original_dict['TDD_sets'][key]['Documents'][document][
                    'Version']
                _temp_dict['ED_status'] = None
                _temp_dict['Changed_sheets'] = \
                original_dict['TDD_sets'][key]['Documents'][document]['Sheets']
                _temp_dict['AMX_AM'] = \
                original_dict['TDD_sets'][key]['Documents'][document][
                    'AMX'] if 'AMX' in \
                              original_dict['TDD_sets'][key]['Documents'][
                                  document] else None
                _temp_dict['Descr_of_change'] = \
                original_dict['TDD_sets'][key]['Documents'][document][
                    'Description']
                _temp_dict['New_rev?'] = None
                normalized_dict['TDD'].append(_temp_dict)

        for key in original_dict['SSC']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Changed'
            _temp_dict['List_SSC'] = key
            _temp_dict['Name_KKS'] = None
            _temp_dict['Descr_of_change_KKS'] = None
            normalized_dict['SSC'].append(_temp_dict)

        for key in original_dict['Supp_descr_docs']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Supporting'
            _temp_dict['File_name'] = key
            _temp_dict['File_content'] = original_dict['Supp_descr_docs'][key]
            normalized_dict['Support_files'].append(_temp_dict)

        return normalized_dict


def open_dict(obj, ini_str):
    for key in obj:
        if type(obj[key]) is dict:
            print(f"{ini_str * '   '}{key}:")
            open_dict(obj[key], ini_str + 1)
        else:
            print(f"{ini_str * '   '}{key}:   {obj[key]}")


def opener(dictus, spc, form='xml'):
    filler = '    '
    if form == 'xml':
        for key in dictus:
            if type(dictus[key]) != dict:
                print(f'{spc * filler}<{key}>')
                print(f'{(spc + 1) * filler}{dictus[key]}')
            else:
                print(f'{spc * filler}<{key}>')
                opener(dictus[key], spc + 1)
            print(f'{spc * filler}</{key}>')
    else:
        def js_ser(obj):
            if isinstance(obj, datetime.datetime):
                return obj.strftime('%d.%m.%Y')
            else:
                return None

        json_str = json.dumps(dictus, default=js_ser, ensure_ascii=False,
                              indent=2)
        return json_str


def from_diff_to_union_excel(normalized_dict):
    wb = smart_load_workbook('D:\\!Digital_twin\\!CR\\CR_parser/template.xlsx')
    ws = wb.worksheets[0]
    ws['B3'].value = normalized_dict['Change_request_No']
    ws['D3'].value = normalized_dict['Reg_date']
    ws['F3'].value = normalized_dict['Contr_change_coord']
    ws['B4'].value = normalized_dict['Type_of_changes']
    ws['D4'].value = normalized_dict['Constr_facility']
    ws['F4'].value = normalized_dict['Type_of_changed_doc']
    ws['H4'].value = normalized_dict['E-log?']
    ws['B6'].value = normalized_dict['Ini_org']
    ws['D6'].value = normalized_dict['Change_ini']
    ws['F6'].value = normalized_dict['Ini_internal_CR']
    ws['B8'].value = normalized_dict['Cod_of_reason']
    ws['D8'].value = normalized_dict['Other_reason']
    ws['B10'].value = normalized_dict['Descr_of_change']
    ws['B11'].value = normalized_dict['Rel_to_prev_CR']
    ws['B13'].value = normalized_dict['Approval_method']
    ws['D13'].value = normalized_dict['Just_simple']
    ws['B16'].value = normalized_dict['NSC_category']
    ws['D16'].value = normalized_dict['Impact_direct_123']
    ws['F16'].value = normalized_dict['Impact_DSA']
    ws['H16'].value = normalized_dict['Impact_structural_geom']
    ws['B17'].value = normalized_dict['Impact_nucl']
    ws['D17'].value = normalized_dict['Impact_fire']
    ws['F17'].value = normalized_dict['Impact_industrial']
    ws['H17'].value = normalized_dict['Impact_environment']
    ws['B18'].value = normalized_dict['Impact_TDD']
    ws['D18'].value = normalized_dict['Impact_lic_doc']
    ws['F18'].value = normalized_dict['Prompt_TDD?']
    ws['H18'].value = normalized_dict['Material_equivalent?']
    ws['B19'].value = normalized_dict['Comments_eng_eval']
    ws['B21'].value = normalized_dict['Impact_contract']
    ws['D21'].value = normalized_dict['Impact_cost']
    ws['F21'].value = normalized_dict['Impact_schedule']
    ws['B22'].value = normalized_dict['Comments_nont_ass']

    ws = wb.create_sheet('Signature list')
    ws.merge_cells('A1:G1')
    ws['A1'].value = 'CR / FCR signature list'
    ws['A2'].value = 'Role /\n Роль'
    ws['B2'].value = 'Position /\n Должность'
    ws['C2'].value = 'Name and Surname /\n Имя и Фамилия'
    ws['D2'].value = 'FMV number /\n Номер лицензии FMV'
    ws['E2'].value = 'HAEA registration /\n Регистрация HAEA'
    ws['F2'].value = 'Signature /\n Подпись'
    ws['G2'].value = 'Date /\n Дата'
    letter_massive = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    formating = excel_formating(ws)
    formating.sheet_view(1)
    formating.col_width(letter_massive, 34)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 8)]
    for cell in _temp_list_of_cells:
        formating.text_alignment(cell, 'center', 'center')
        if cell == [1, 1]:
            formating.text_font(cell, text_size=14)
        else:
            formating.text_font(cell, text_size=14, text_bold=True)

    _temp_counter = 2
    for key in normalized_dict['Signature_list']:
        _temp_counter += 1
        for letter in letter_massive:
            formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                     'center')
        ws[f'A{_temp_counter}'].value = key['Role']
        ws[f'B{_temp_counter}'].value = key['Position']
        ws[f'C{_temp_counter}'].value = key['Name_sur']
        ws[f'D{_temp_counter}'].value = key['FMV_number']
        ws[f'E{_temp_counter}'].value = key['HAEA_reg']
        ws[f'F{_temp_counter}'].value = None
        if isinstance(key['Date'], datetime.datetime):
            str_date = key['Date'].strftime('%d.%m.%Y')
        else:
            str_date = key['Date']
        ws[f'G{_temp_counter}'].value = str_date
    formating.borders('thin')

    ws = wb.create_sheet('4&6 Affected TDD')
    ws.merge_cells('A1:N1')
    ws['A1'].value = '4.1. Changed TDD'
    ws['A2'].value = 'Document set\ncode: /\nКод комплекта\nдокументов:'
    ws[
        'B2'].value = 'Document set\nname: /\nНаименование\nкомплекта\nдокументов:'
    ws[
        'C2'].value = 'Document set\nRevision: /\nРевизия\nкомплекта\nдокументов:'
    ws['D2'].value = 'Document set\nVersion: /\nВерсия комплекта\nдокументов:'
    ws['E2'].value = 'Document set\nStatus: /\nСтатус комплекта\nдокументов:'
    ws['F2'].value = 'ED set Code*: /\nКод выпускаемого\nдокумента*:'
    ws['G2'].value = 'ED set Name*: /\nНаименование\nвыпускаемого\nдокумента*:'
    ws['H2'].value = 'ED set Revision*: /\nРевизия\nвыпускаемого\nдокумента*:'
    ws['I2'].value = 'ED set Version*: /\nВерсия\nвыпускаемого\nдокумента*:'
    ws['J2'].value = 'ED set Status*: /\nСтатус\nвыпускаемого\nдокумента*:'
    ws['K2'].value = 'Changed sheets: /\nИзмененные\nлисты:'
    ws['L2'].value = 'Amx/Change\nversion: /\nНомер\nAM/Изменения:'
    ws['M2'].value = 'Change\ndescription: /\nОписание\nизменения:'
    ws[
        'N2'].value = 'New revision\nrequired: /\nТребуется ли\nвыпуск новой\nревизии:'

    letter_massive = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                      'L', 'M', 'N']
    formating = excel_formating(ws)
    formating.sheet_view(1)
    formating.col_width(letter_massive, 17)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 15)]
    for cell in _temp_list_of_cells:
        if cell == [1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    _temp_counter = 2
    for key in normalized_dict['TDD']:
        if key['Type'] == 'TDD':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['Set_code']
            ws[f'B{_temp_counter}'].value = key['Set_name']
            ws[f'C{_temp_counter}'].value = key['Set_rev']
            ws[f'D{_temp_counter}'].value = key['Ser_version']
            ws[f'E{_temp_counter}'].value = key['Set_status']
            ws[f'F{_temp_counter}'].value = key['ED_code']
            ws[f'G{_temp_counter}'].value = key['ED_name']
            ws[f'H{_temp_counter}'].value = key['ED_rev']
            ws[f'I{_temp_counter}'].value = key['ED_version']
            ws[f'J{_temp_counter}'].value = key['ED_status']
            ws[f'K{_temp_counter}'].value = key['Changed_sheets']
            ws[f'L{_temp_counter}'].value = key['AMX_AM']
            ws[f'M{_temp_counter}'].value = key['Descr_of_change']
            ws[f'N{_temp_counter}'].value = key['New_rev?']

    _temp_counter += 1
    ws.merge_cells(f'A{_temp_counter}:N{_temp_counter}')
    ws[f'A{_temp_counter}'] = '6.3. Affected TDD'
    _temp_counter += 1
    for letter in letter_massive:
        ws[f'{letter}{_temp_counter}'].value = ws[f'{letter}2'].value
    formating.color([_temp_counter - 1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[_temp_counter - 1, 1]] + [[_temp_counter, x] for x
                                                      in range(1, 15)]
    for cell in _temp_list_of_cells:
        if cell == [_temp_counter - 1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    for key in normalized_dict['TDD']:
        if key['Type'] == 'Configur':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['Set_code']
            ws[f'B{_temp_counter}'].value = key['Set_name']
            ws[f'C{_temp_counter}'].value = key['Set_rev']
            ws[f'D{_temp_counter}'].value = key['Ser_version']
            ws[f'E{_temp_counter}'].value = key['Set_status']
            ws[f'F{_temp_counter}'].value = key['ED_code']
            ws[f'G{_temp_counter}'].value = key['ED_name']
            ws[f'H{_temp_counter}'].value = key['ED_rev']
            ws[f'I{_temp_counter}'].value = key['ED_version']
            ws[f'J{_temp_counter}'].value = key['ED_status']
            ws[f'K{_temp_counter}'].value = key['Changed_sheets']
            ws[f'L{_temp_counter}'].value = key['AMX_AM']
            ws[f'M{_temp_counter}'].value = key['Descr_of_change']
            ws[f'N{_temp_counter}'].value = key['New_rev?']
    formating.borders('thin')

    ws = wb.create_sheet('4&6 Changed SSC')
    ws.merge_cells('A1:C1')
    ws['A1'].value = '4.2. Changed SSC'
    ws['A2'].value = 'List of affected SSC: /\nПеречень затронутых KKS:'
    ws['B2'].value = 'Name of affected KKS: / \nНаименование затронутого KKS:'
    ws['C2'].value = 'Description of KKS change: /\nОписание изменения KKS:'
    formating = excel_formating(ws)
    formating.sheet_view(1)
    letter_massive = ['A', 'B', 'C']
    formating.col_width(['A'], 30)
    formating.col_width(['B'], 47)
    formating.col_width(['C'], 54)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 4)]
    for cell in _temp_list_of_cells:
        if cell == [1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    _temp_counter = 2
    for key in normalized_dict['SSC']:
        if key['Type'] == 'Changed':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['List_SSC']
            ws[f'B{_temp_counter}'].value = key['Name_KKS']
            ws[f'C{_temp_counter}'].value = key['Descr_of_change_KKS']

    _temp_counter += 1
    ws.merge_cells(f'A{_temp_counter}:C{_temp_counter}')
    ws[f'A{_temp_counter}'] = '6.4. Impacted SSC'
    _temp_counter += 1
    for letter in letter_massive:
        ws[f'{letter}{_temp_counter}'].value = ws[f'{letter}2'].value
    formating.color([_temp_counter - 1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[_temp_counter - 1, 1]] + [[_temp_counter, x] for x
                                                      in range(1, 4)]
    for cell in _temp_list_of_cells:
        if cell == [_temp_counter - 1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    for key in normalized_dict['SSC']:
        if key['Type'] == 'Configur':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['List_SSC']
            ws[f'B{_temp_counter}'].value = key['Name_KKS']
            ws[f'C{_temp_counter}'].value = key['Descr_of_change_KKS']
    formating.borders('thin')

    ws = wb.create_sheet('3&7 Supporting Files')
    ws.merge_cells('A1:B1')
    ws['A1'].value = '3.1. Supporting and describing documents'
    ws['A2'].value = 'File name: /\nИмя файла:'
    ws['B2'].value = 'File content: / \nСодержание файла:'
    formating = excel_formating(ws)
    formating.sheet_view(1)
    letter_massive = ['A', 'B']
    formating.col_width(['A'], 62)
    formating.col_width(['B'], 56)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 3)]
    for cell in _temp_list_of_cells:
        if cell == [1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    _temp_counter = 2
    for key in normalized_dict['Support_files']:
        if key['Type'] == 'Supporting':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['File_name']
            ws[f'B{_temp_counter}'].value = key['File_content']

    _temp_counter += 1
    ws.merge_cells(f'A{_temp_counter}:B{_temp_counter}')
    ws[f'A{_temp_counter}'] = '7.1. Documents justifying the decision'
    _temp_counter += 1
    for letter in letter_massive:
        ws[f'{letter}{_temp_counter}'].value = ws[f'{letter}2'].value
    formating.color([_temp_counter - 1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[_temp_counter - 1, 1]] + [[_temp_counter, x] for x
                                                      in range(1, 3)]
    for cell in _temp_list_of_cells:
        if cell == [_temp_counter - 1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    for key in normalized_dict['Support_files']:
        if key['Type'] == 'Justifying':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['File_name']
            ws[f'B{_temp_counter}'].value = key['File_content']
    formating.borders('thin')

    file_name_list = normalized_dict['File_name'].split('.')
    file_name = '.'.join(file_name_list[:-1]) + '_NS.' + file_name_list[-1]
    # NOTE: ensure excel_path is globally defined or passed, adjusting path logic as in original
    output_dir = os.path.join(os.path.dirname(file_name), 'output')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    wb.save(os.path.join(output_dir, file_name.split('\\')[-1]))
    return normalized_dict['File_name']


class excel_formating:
    def __init__(self, ws):
        self.ws = ws

    def col_width(self, lett_list: list, value):
        for letter in lett_list:
            self.ws.column_dimensions[letter].width = value

    def text_alignment(self, place: list or str, hor, vert):
        if type(place) is list:
            cell = self.ws.cell(*place)
        else:
            cell = self.ws[place]
        cell.alignment = Alignment(horizontal=hor, vertical=vert,
                                   wrapText=True)

    def color(self, place: list, start_col: str, end_col: str):
        cell = self.ws.cell(*place)
        filling = PatternFill(start_color=start_col, end_color=end_col,
                              fill_type='solid')
        cell.fill = filling

    def text_font(self, place: list, text_size: int, text_bold=False):
        cell = self.ws.cell(*place)
        cell.font = Font(bold=text_bold, size=text_size)

    def sheet_view(self, page_number):
        self.ws.sheet_view.view = 'pageBreakPreview'
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToWidth = page_number
        self.ws.page_setup.fitToHeight = False

    def borders(self, border_type_name):
        border_type = Border(
            left=Side(border_style=border_type_name, color='000000'),
            right=Side(border_style=border_type_name, color='000000'),
            top=Side(border_style=border_type_name, color='000000'),
            bottom=Side(border_style=border_type_name, color='000000'))
        for row in range(1, self.ws.max_row + 1):
            for col in range(1, self.ws.max_column + 1):
                self.ws.cell(row=row, column=col).border = border_type


def output(option):
    global result  # Ensuring result is accessible by dicts_normalization
    if option == 0:
        excel_path = 'D:\\!Digital_twin\\!CR\\CR_parser\\cr_test\\18_05\\'
    else:
        if option == 'pwd':
            excel_path = ''
        else:
            excel_path = input(
                'Please, enter the FULL path to the directory with CR, CR.D, FCR: ')
            if excel_path and excel_path[-1] != '\\':
                excel_path += '\\'

    list_of_tables = os.listdir(path=excel_path if excel_path else '.')
    list_of_tables = list(filter(
        lambda x: 'xls' in x.split('.')[-1] and x[-1] != 'b' and ('CR' in x),
        list_of_tables))
    result = {}

    for file in list_of_tables:
        full_path = f'{excel_path}{file}' if excel_path else file
        result[file] = main_func(full_path)

    if option == 0:
        CR_list = list(filter(lambda x: '?FCR.D-' in x, list(result.keys())))
        for CR in CR_list:
            a = dicts_normalization(CR)
            from_diff_to_union_excel(a)
    else:
        for dictus in result:
            if dictus[:4] == 'CR.D' and dictus in result and result[dictus]:
                for code in result[dictus].get('TDD', {}):
                    if 'Impact' in result[dictus]['TDD'][code] and isinstance(
                            result[dictus]['TDD'][code]['Impact'], int):
                        bin_impact = f"{result[dictus]['TDD'][code]['Impact']:05b}"
                        bin_impact = [True if int(i) else False for i in
                                      bin_impact]
                        result[dictus]['TDD'][code]['Impact'] = {
                            'NS': bin_impact[0], 'FS': bin_impact[1],
                            'IS': bin_impact[2], 'ES': bin_impact[3],
                            'SS': bin_impact[4]}

            if result[dictus]:
                short_name = '.'.join(dictus.split('.')[:-1])
                file_name = excel_path + short_name + '.txt' if excel_path else short_name + '.txt'
                with open(file_name, 'wb+') as out_json:
                    out_json.write(b'\xff\xfe')
                    out_json.write(
                        opener(result[dictus], 0, form='json').encode(
                            'utf-16-le'))
                    print(f'{short_name}.txt is ready!')


if __name__ == '__main__':
    output('pwd')
    # os.system('pause') # Убрал, так как на Mac это вызовет ошибку. Для Windows можете раскомментировать.